#!/usr/bin/env python3
import sqlite3
import openpyxl
from datetime import datetime
import os

def create_database():
    """Create SQLite database with six tables as specified"""
    
    # Connect to SQLite database (creates if doesn't exist)
    conn = sqlite3.connect('medical_lock_hospitals.db')
    cursor = conn.cursor()
    
    # Create Documents table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            doc_id TEXT PRIMARY KEY,
            source_name TEXT,
            type TEXT,
            link TEXT,
            notes TEXT
        )
    ''')
    
    # Create Stations table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stations (
            station_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            region TEXT,
            country TEXT,
            latitude REAL,
            longitude REAL,
            notes TEXT
        )
    ''')
    
    # Create Station Reports table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS station_reports (
            report_id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id TEXT,
            station_id INTEGER,
            FOREIGN KEY (doc_id) REFERENCES documents (doc_id),
            FOREIGN KEY (station_id) REFERENCES stations (station_id)
        )
    ''')
    
    # Create Women Data table (from women_admission sheet)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS women_data (
            unique_id TEXT PRIMARY KEY,
            doc_id TEXT,
            source_name TEXT,
            source_type TEXT,
            region TEXT,
            station TEXT,
            country TEXT,
            year INTEGER,
            women_start_register INTEGER,
            women_added INTEGER,
            FOREIGN KEY (doc_id) REFERENCES documents (doc_id)
        )
    ''')
    
    # Create Troop Data table (from troops_admission sheet)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS troop_data (
            unique_id TEXT PRIMARY KEY,
            doc_id TEXT,
            source_name TEXT,
            source_type TEXT,
            region TEXT,
            station TEXT,
            country TEXT,
            year INTEGER,
            regiments TEXT,
            avg_strength REAL,
            FOREIGN KEY (doc_id) REFERENCES documents (doc_id)
        )
    ''')
    
    # Create Hospital Operations table (from Hospitals sheet)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS hospital_operations (
            hid TEXT PRIMARY KEY,
            doc_id TEXT,
            source_name TEXT,
            source_type TEXT,
            year INTEGER,
            region TEXT,
            station TEXT,
            country TEXT,
            act TEXT,
            class TEXT,
            staff_medical_officers INTEGER,
            staff_hospital_assistants INTEGER,
            staff_matron INTEGER,
            staff_coolies INTEGER,
            staff_peons INTEGER,
            staff_watermen INTEGER,
            ops_inspection_regularity TEXT,
            ops_unlicensed_control_notes TEXT,
            ops_committee_activity_notes TEXT,
            FOREIGN KEY (doc_id) REFERENCES documents (doc_id)
        )
    ''')
    
    conn.commit()
    print("Database schema created successfully!")
    return conn, cursor

def extract_data_from_excel(conn, cursor):
    """Extract data from Excel file and populate database tables"""
    
    # Load the workbook
    workbook = openpyxl.load_workbook('DS_Dataset.xlsx', data_only=True)
    
    # Extract unique documents first
    documents = set()
    stations = set()
    
    # Process each sheet to collect unique documents and stations
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Get headers
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
                
            # Extract document info (assuming doc_id is in column 1, source_name in column 2, etc.)
            if len(row) > 1 and row[1]:  # doc_id
                doc_id = str(row[1])
                source_name = str(row[2]) if len(row) > 2 and row[2] else ""
                source_type = str(row[3]) if len(row) > 3 and row[3] else ""
                documents.add((doc_id, source_name, source_type))
            
            # Extract station info
            if len(row) > 6 and row[6]:  # station name
                station_name = str(row[6])
                region = str(row[5]) if len(row) > 5 and row[5] else ""
                country = str(row[7]) if len(row) > 7 and row[7] else ""
                stations.add((station_name, region, country))
    
    # Insert documents
    print("Inserting documents...")
    for doc_id, source_name, source_type in documents:
        cursor.execute('''
            INSERT OR IGNORE INTO documents (doc_id, source_name, type)
            VALUES (?, ?, ?)
        ''', (doc_id, source_name, source_type))
    
    # Insert stations
    print("Inserting stations...")
    for station_name, region, country in stations:
        cursor.execute('''
            INSERT OR IGNORE INTO stations (name, region, country)
            VALUES (?, ?, ?)
        ''', (station_name, region, country))
    
    conn.commit()
    
    # Process women_admission sheet
    print("Processing women_admission data...")
    women_sheet = workbook['women_admission']
    headers = [cell.value for cell in women_sheet[1]]
    
    for row in women_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
            
        cursor.execute('''
            INSERT OR REPLACE INTO women_data 
            (unique_id, doc_id, source_name, source_type, region, station, country, year, women_start_register, women_added)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            str(row[0]) if row[0] else None,  # unique_id
            str(row[1]) if row[1] else None,  # doc_id
            str(row[2]) if row[2] else None,  # source_name
            str(row[3]) if row[3] else None,  # source_type
            str(row[4]) if row[4] else None,  # region
            str(row[5]) if row[5] else None,  # station
            str(row[6]) if row[6] else None,  # country
            int(row[7]) if row[7] and str(row[7]).isdigit() else None,  # year
            int(row[8]) if row[8] and str(row[8]).isdigit() else None,  # women_start_register
            int(row[9]) if row[9] and str(row[9]).isdigit() else None   # women_added
        ))
    
    # Process troops_admission sheet
    print("Processing troops_admission data...")
    troops_sheet = workbook['troops_admission']
    
    for row in troops_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
            
        cursor.execute('''
            INSERT OR REPLACE INTO troop_data 
            (unique_id, doc_id, source_name, source_type, region, station, country, year, regiments, avg_strength)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            str(row[0]) if row[0] else None,  # unique_id
            str(row[1]) if row[1] else None,  # doc_id
            str(row[2]) if row[2] else None,  # source_name
            str(row[3]) if row[3] else None,  # source_type
            str(row[4]) if row[4] else None,  # region
            str(row[5]) if row[5] else None,  # station
            str(row[6]) if row[6] else None,  # country
            int(row[7]) if row[7] and str(row[7]).isdigit() else None,  # year
            str(row[8]) if row[8] else None,  # regiments
            float(row[9]) if row[9] and str(row[9]).replace('.', '').isdigit() else None  # avg_strength
        ))
    
    # Process Hospitals sheet
    print("Processing Hospitals data...")
    hospitals_sheet = workbook['Hospitals']
    
    for row in hospitals_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
            
        cursor.execute('''
            INSERT OR REPLACE INTO hospital_operations 
            (hid, doc_id, source_name, source_type, year, region, station, country, act, class,
             staff_medical_officers, staff_hospital_assistants, staff_matron, staff_coolies,
             staff_peons, staff_watermen, ops_inspection_regularity, ops_unlicensed_control_notes,
             ops_committee_activity_notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            str(row[0]) if row[0] else None,  # hid
            str(row[1]) if row[1] else None,  # doc_id
            str(row[2]) if row[2] else None,  # source_name
            str(row[3]) if row[3] else None,  # source_type
            int(row[4]) if row[4] and str(row[4]).isdigit() else None,  # year
            str(row[5]) if row[5] else None,  # region
            str(row[6]) if row[6] else None,  # station
            str(row[7]) if row[7] else None,  # country
            str(row[8]) if row[8] else None,  # act
            str(row[9]) if row[9] else None,  # class
            int(row[10]) if row[10] and str(row[10]).isdigit() else None,  # staff_medical_officers
            int(row[11]) if row[11] and str(row[11]).isdigit() else None,  # staff_hospital_assistants
            int(row[12]) if row[12] and str(row[12]).isdigit() else None,  # staff_matron
            int(row[13]) if row[13] and str(row[13]).isdigit() else None,  # staff_coolies
            int(row[14]) if row[14] and str(row[14]).isdigit() else None,  # staff_peons
            int(row[15]) if row[15] and str(row[15]).isdigit() else None,  # staff_watermen
            str(row[16]) if row[16] else None,  # ops_inspection_regularity
            str(row[17]) if row[17] else None,  # ops_unlicensed_control_notes
            str(row[18]) if row[18] else None   # ops_committee_activity_notes
        ))
    
    # Create station reports relationships
    print("Creating station reports relationships...")
    cursor.execute('''
        INSERT INTO station_reports (doc_id, station_id)
        SELECT DISTINCT w.doc_id, s.station_id
        FROM women_data w
        JOIN stations s ON w.station = s.name
        WHERE w.doc_id IS NOT NULL AND s.station_id IS NOT NULL
    ''')
    
    cursor.execute('''
        INSERT OR IGNORE INTO station_reports (doc_id, station_id)
        SELECT DISTINCT t.doc_id, s.station_id
        FROM troop_data t
        JOIN stations s ON t.station = s.name
        WHERE t.doc_id IS NOT NULL AND s.station_id IS NOT NULL
    ''')
    
    cursor.execute('''
        INSERT OR IGNORE INTO station_reports (doc_id, station_id)
        SELECT DISTINCT h.doc_id, s.station_id
        FROM hospital_operations h
        JOIN stations s ON h.station = s.name
        WHERE h.doc_id IS NOT NULL AND s.station_id IS NOT NULL
    ''')
    
    conn.commit()
    print("Data extraction completed!")

def verify_database(cursor):
    """Verify database structure and data integrity"""
    print("\n=== Database Verification ===")
    
    # Check table counts
    tables = ['documents', 'stations', 'station_reports', 'women_data', 'troop_data', 'hospital_operations']
    
    for table in tables:
        cursor.execute(f'SELECT COUNT(*) FROM {table}')
        count = cursor.fetchone()[0]
        print(f"{table}: {count} records")
    
    # Show sample data from each table
    print("\n=== Sample Data ===")
    for table in tables:
        print(f"\n--- {table} (first 3 records) ---")
        cursor.execute(f'SELECT * FROM {table} LIMIT 3')
        rows = cursor.fetchall()
        for row in rows:
            print(row)

if __name__ == "__main__":
    print("Creating medical lock hospitals database...")
    
    # Create database and schema
    conn, cursor = create_database()
    
    # Extract and populate data
    extract_data_from_excel(conn, cursor)
    
    # Verify database
    verify_database(cursor)
    
    # Close connection
    conn.close()
    
    print(f"\nDatabase created successfully: medical_lock_hospitals.db")
    print("Database contains 6 tables as requested:")
    print("1. documents (doc_id, source_name, type, link, notes)")
    print("2. stations (station_id, name, region, country, lat/lon, notes)")
    print("3. station_reports (report_id, doc_id, station_id)")
    print("4. women_data (from women_admission sheet)")
    print("5. troop_data (from troops_admission sheet)")
    print("6. hospital_operations (from Hospitals sheet)")
